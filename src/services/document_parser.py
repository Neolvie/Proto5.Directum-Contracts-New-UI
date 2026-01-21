from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List

import mammoth
from openpyxl import load_workbook
from pypdf import PdfReader


@dataclass
class ParsedDocument:
    name: str
    text: str
    pages: List[str]


def parse_document(file_path: Path, display_name: str) -> ParsedDocument:
    """Распарсить документ в зависимости от расширения."""
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        pages = extract_pdf_text(file_path)
    elif suffix in {".docx", ".doc"}:
        pages = [extract_docx_text(file_path)]
    elif suffix in {".xlsx", ".xls"}:
        pages = [extract_xlsx_text(file_path)]
    elif suffix in {".md", ".txt"}:
        pages = [extract_md_text(file_path)]
    else:
        pages = [extract_md_text(file_path)]

    combined = normalize_text(pages)
    return ParsedDocument(name=display_name, text=combined, pages=pages)


def extract_docx_text(file_path: Path) -> str:
    """Извлечь текст из DOCX с помощью mammoth."""
    result = mammoth.extract_raw_text(file_path)
    return result.value.strip()


def extract_pdf_text(file_path: Path) -> List[str]:
    """Извлечь текст по страницам PDF."""
    reader = PdfReader(str(file_path))
    pages: List[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        pages.append(text.strip())
    return pages


def extract_xlsx_text(file_path: Path) -> str:
    """Извлечь текст из XLSX по всем листам."""
    workbook = load_workbook(file_path, data_only=True)
    rows: List[str] = []
    for sheet in workbook.worksheets:
        rows.append(f"Лист: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell) for cell in row if cell is not None]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows).strip()


def extract_md_text(file_path: Path) -> str:
    """Прочитать текстовый файл как UTF-8."""
    return file_path.read_text(encoding="utf-8", errors="ignore").strip()


def normalize_text(pages: List[str], max_chars: int = 120_000) -> str:
    """Собрать текст с маркерами страниц и ограничить размер."""
    formatted_pages = []
    for idx, page_text in enumerate(pages, start=1):
        marker = f"=== PAGE {idx} ==="
        formatted_pages.append(f"{marker}\n{page_text}".strip())
    combined = "\n\n".join(formatted_pages).strip()
    if len(combined) <= max_chars:
        return combined
    return combined[: max_chars - 200] + "\n\n[Текст усечен]"
